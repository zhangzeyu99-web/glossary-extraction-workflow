from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

import extract_glossary as extractor


def import_glossary(
    workbook_path: Path,
    *,
    sheet_name: str,
    curated_rules_path: Path | None,
) -> tuple[int, Path | None]:
    curated_rules = extractor.load_curated_rules(curated_rules_path)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(min_row=1, values_only=True)
    headers = list(next(rows))
    id_index = extractor.resolve_column_index(headers, "ID")
    cn_index = extractor.resolve_column_index(headers, "CN")
    en_index = extractor.resolve_column_index(headers, "EN")
    en2_index = extractor.resolve_column_index(headers, "EN2")

    imported_count = 0
    for row in rows:
        _row_id = "" if row[id_index] is None else str(row[id_index])
        cn = extractor.clean_text(row[cn_index])
        en = extractor.clean_text(row[en_index])
        en2 = extractor.clean_text(row[en2_index])
        if not cn or not en:
            continue
        state = extractor.get_curated_term_state(curated_rules, cn)
        state["approved_en"] = en
        state["approved_en2"] = en2
        state["block_en2"] = not bool(en2)
        imported_count += 1

    workbook.close()
    extractor.save_curated_rules(curated_rules_path, curated_rules)
    return imported_count, curated_rules_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import a final glossary workbook back into curated glossary rules.")
    parser.add_argument("input_path", help="Path to the final glossary workbook.")
    parser.add_argument("--sheet", default="Glossary", help="Worksheet name. Default: Glossary")
    parser.add_argument(
        "--curated-rules",
        default=str(extractor.DEFAULT_CURATED_RULES),
        help="Path to the curated glossary rules JSON file. Default: data/experience/curated_terms.json",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    workbook_path = Path(args.input_path)
    curated_rules_path = Path(args.curated_rules) if args.curated_rules else None
    imported_count, saved_path = import_glossary(
        workbook_path,
        sheet_name=args.sheet,
        curated_rules_path=curated_rules_path,
    )
    print(f"INPUT={workbook_path}")
    print(f"SHEET={args.sheet}")
    print(f"IMPORTED_TERMS={imported_count}")
    print(f"CURATED_RULES={saved_path or 'disabled'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
