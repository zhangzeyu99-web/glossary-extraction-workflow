from __future__ import annotations

import argparse
import hashlib
import html
import json
import posixpath
import re
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CURATED_RULES = REPO_ROOT / "data" / "experience" / "curated_terms.json"
DEFAULT_OBSERVATIONS_STORE = REPO_ROOT / "data" / "experience" / "observed_terms.json"
DEFAULT_LEGACY_EXPERIENCE_STORE = REPO_ROOT / "data" / "experience" / "term_memory.json"
CURATED_VERSION = 1
OBSERVATION_VERSION = 1

SENTENCE_PUNCT_RE = re.compile(r"[，。！？；：,.!?;:\n]")
CJK_RE = re.compile(r"[\u4e00-\u9fff]")
PLACEHOLDER_RE = re.compile(r"\{\d+\}|%[sd]|\\n")
BRACKET_TAG_RE = re.compile(r"\[[^\]]*\]")
HTML_TAG_RE = re.compile(r"<[^>]+>")
SPACE_RE = re.compile(r"\s+")
NON_TERM_RE = re.compile(r"^[\W_]+$|^[IVXⅠⅡⅢⅣⅤⅥⅦⅧⅨⅩ]+$")
CAMEL_SPLIT_RE = re.compile(r"(?<=[a-z])(?=[A-Z])")
EN_COMPARE_RE = re.compile(r"[^a-z0-9+ ]+")
EN_WORD_RE = re.compile(r"[a-z0-9+]+")

RARITY_TERMS = {
    "普通",
    "精良",
    "精英",
    "卓越",
    "史诗",
    "传说",
    "神话",
    "高级",
}
RESOURCE_TERMS = {
    "钻石",
    "石油",
    "体力",
    "宝石",
    "积分",
    "碎片",
    "材料",
    "芯片",
    "能量",
    "金币",
    "经验",
    "奖励",
    "礼包",
}
STAT_TERMS = {
    "攻击",
    "攻击力",
    "防御",
    "生命",
    "伤害",
    "伤害+",
    "暴击",
    "暴击伤害",
    "闪电",
    "火焰",
    "冰霜",
    "物理",
    "闪电伤害",
    "火焰伤害",
    "冰霜伤害",
    "物理伤害",
    "能量伤害",
}
ACTION_TERMS = {
    "获得",
    "获取",
    "领取",
    "使用",
    "合成",
    "升级",
    "强化",
    "突破",
    "升星",
    "激活",
    "解锁",
    "购买",
    "报名",
    "兑换",
    "刷新",
    "前往",
    "开始",
    "加入",
    "退出",
    "上阵",
    "建造",
    "邀请",
    "重置",
    "探索",
    "挑战",
}
SYSTEM_TERMS = {
    "公会",
    "竞技场",
    "战令",
    "签到",
    "商城",
    "商店",
    "背包",
    "排行",
    "排行榜",
    "活动",
    "防御塔",
    "兵营",
    "基地",
    "据点",
    "任务",
}
OBJECT_TERMS = {
    "英雄",
    "技能",
    "装备",
    "建筑",
    "武器",
    "士兵",
    "坐骑",
    "收藏品",
    "头像",
    "好友",
    "道具",
}
STATUS_TERMS = {
    "当前",
    "暂无",
    "不足",
    "已满",
    "失败",
    "可领取",
    "拥有",
    "最大",
    "剩余",
    "排名",
    "段位",
}

HIGH_CONFUSION_TERMS = (
    RARITY_TERMS
    | RESOURCE_TERMS
    | ACTION_TERMS
    | STATUS_TERMS
    | {
        "战力",
        "品质",
        "等级",
        "积分",
        "攻击",
        "觉醒",
        "奖励",
        "选择",
        "强化",
        "合成",
        "突破",
        "推荐",
    }
)

PROJECT_SIGNAL_GROUPS = {
    "合成/经营": {
        "合成",
        "订单",
        "生产",
        "生产机",
        "生产机器",
        "生成器",
        "仓库",
        "菜品",
        "烹饪",
        "顾客",
        "Merge",
    },
    "花店/装修": {
        "花店",
        "花束",
        "鲜花",
        "玫瑰",
        "百合",
        "装饰",
        "装修",
        "修复",
        "翻新",
        "花园",
        "Florist",
    },
    "休闲/女性向": {
        "可爱",
        "漂亮",
        "温馨",
        "甜",
        "咖啡",
        "甜品",
        "裙",
        "珠宝",
        "约会",
        "浪漫",
        "美女",
        "小姐",
    },
    "战斗/RPG养成": {
        "战斗",
        "攻击",
        "防御",
        "生命",
        "伤害",
        "暴击",
        "技能",
        "英雄",
        "装备",
        "武器",
        "首领",
        "BOSS",
        "怪物",
        "关卡",
        "挑战",
    },
    "基地/建筑经营": {
        "基地",
        "建筑",
        "兵营",
        "防御塔",
        "建造",
        "升级",
        "营地",
        "总部",
        "据点",
        "采集",
        "生产",
    },
    "活动/商业化": {
        "活动",
        "签到",
        "战令",
        "礼包",
        "充值",
        "商店",
        "商城",
        "购买",
        "限时",
        "奖励",
        "抽奖",
        "召唤",
    },
    "社交/公会竞争": {
        "公会",
        "联盟",
        "好友",
        "排行榜",
        "排名",
        "竞技场",
        "聊天",
        "邀请",
        "成员",
        "队伍",
    },
    "飞行/射击题材": {
        "飞机",
        "战机",
        "飞行员",
        "机库",
        "导弹",
        "空袭",
        "射击",
        "僚机",
        "弹幕",
    },
    "末日/生存题材": {
        "幸存者",
        "僵尸",
        "末日",
        "避难所",
        "感染",
        "生存",
        "废土",
        "救援",
    },
    "剧情/叙事": {
        "剧情",
        "章节",
        "对话",
        "故事",
        "任务",
        "探索",
        "冒险",
        "线索",
        "选择",
        "先生",
        "老板",
        "小姐",
        "等等",
        "拜托",
    },
}

CATEGORY_LABELS = {
    "rarity": "稀有度/品质",
    "resource": "资源/货币/奖励",
    "stat": "战斗属性/数值",
    "action": "UI操作动词",
    "system": "系统/玩法名",
    "object": "角色/装备/对象",
    "status": "状态/进度",
    "other": "其他",
}


@dataclass
class Record:
    row_id: str
    source: str
    target: str


def clean_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = html.unescape(text)
    text = HTML_TAG_RE.sub(" ", text)
    text = BRACKET_TAG_RE.sub("", text)
    text = PLACEHOLDER_RE.sub("", text)
    text = SPACE_RE.sub(" ", text).strip()
    return text


def normalize_english_for_compare(text: str) -> str:
    text = clean_text(text)
    text = CAMEL_SPLIT_RE.sub(" ", text)
    text = text.lower()
    text = re.sub(r"\s*\+\s*", "+", text)
    text = re.sub(r"[-_/]+", " ", text)
    text = EN_COMPARE_RE.sub(" ", text)
    text = SPACE_RE.sub(" ", text).strip()

    normalized_tokens: list[str] = []
    for token in text.split():
        if token.endswith("ies") and len(token) > 4:
            token = token[:-3] + "y"
        elif token.endswith(("oes", "ses", "xes", "zes", "ches", "shes")) and len(token) > 4:
            token = token[:-2]
        elif token.endswith("s") and len(token) > 3 and not token.endswith(("ss", "us", "is")):
            token = token[:-1]
        normalized_tokens.append(token)
    return " ".join(normalized_tokens)


def is_same_or_extended_usage(example_en: str, actual_en: str) -> bool:
    example_norm = normalize_english_for_compare(example_en)
    actual_norm = normalize_english_for_compare(actual_en)
    if not actual_norm or not example_norm:
        return False
    if actual_norm == example_norm:
        return True
    return f" {example_norm} " in f" {actual_norm} "


def split_usage_buckets(example_en: str, actual_counter: Counter[str]) -> tuple[Counter[str], Counter[str]]:
    example_counter: Counter[str] = Counter()
    manual_counter: Counter[str] = Counter()
    for actual_en, count in actual_counter.items():
        if is_same_or_extended_usage(example_en=example_en, actual_en=actual_en):
            example_counter[actual_en] += count
        else:
            manual_counter[actual_en] += count
    return example_counter, manual_counter


def collect_translation_diff(example_en: str, actual_counter: Counter[str]) -> dict[str, object]:
    same_counter, diff_counter = split_usage_buckets(
        example_en=example_en,
        actual_counter=actual_counter,
    )
    return {
        "has_diff": "Yes" if diff_counter else "No",
        "same_or_format_only_count": sum(same_counter.values()),
        "diff_count": sum(diff_counter.values()),
        "diff_variants": join_counter(diff_counter, limit=8),
        "diff_type": "manual_adaptation" if diff_counter else "",
    }


def token_roots(text: str) -> list[str]:
    roots: list[str] = []
    for token in EN_WORD_RE.findall(normalize_english_for_compare(text)):
        root = token
        if root.endswith("ing") and len(root) > 5:
            root = root[:-3]
        elif root.endswith("ed") and len(root) > 4:
            root = root[:-2]
        elif root.endswith("er") and len(root) > 4:
            root = root[:-2]
        elif root.endswith("ation") and len(root) > 7:
            root = root[:-5] + "e"
        roots.append(root)
    return roots


def titleize_word(word: str) -> str:
    if word.isupper():
        return word
    if word in {"hp", "atk", "def", "dmg", "cp"}:
        return word.upper()
    return word.capitalize()


def choose_en2_value(
    example_en: str,
    exact_diff_counter: Counter[str],
    manual_counter: Counter[str],
) -> str:
    if exact_diff_counter:
        return " | ".join(text for text, _ in exact_diff_counter.most_common(3))
    if not manual_counter:
        return ""

    manual_variants = manual_counter.most_common()
    top_text, top_count = manual_variants[0]
    second_count = manual_variants[1][1] if len(manual_variants) > 1 else 0
    total = sum(manual_counter.values())
    if top_count >= 2 and top_count > second_count and top_count / total >= 0.45:
        top_norm = normalize_english_for_compare(top_text)
        if top_norm and all(
            normalize_english_for_compare(text) == top_norm
            or is_same_or_extended_usage(example_en=top_text, actual_en=text)
            or is_same_or_extended_usage(example_en=text, actual_en=top_text)
            for text, _count in manual_variants[1:]
        ):
            return top_text

    example_roots = set(token_roots(example_en))
    root_counter: Counter[str] = Counter()

    for text, count in manual_counter.items():
        for root in token_roots(text):
            if root in example_roots or root in {"the", "a", "an", "of", "to", "for", "in", "on", "with", "and"}:
                continue
            root_counter[root] += count

    if not root_counter:
        return ""

    top_root, top_count = root_counter.most_common(1)[0]
    second_count = root_counter.most_common(2)[1][1] if len(root_counter) > 1 else 0
    if top_count < 2 or top_count <= second_count:
        return ""
    if top_count / total < 0.45:
        return ""
    return titleize_word(top_root)


def is_short_usage_candidate(record: Record, term: str, example_en: str) -> bool:
    if not record.target:
        return False
    if record.source == term:
        return True
    source_limit = max(8, len(term) + 4)
    target_limit = max(28, len(example_en) + 12) if example_en else 28
    return len(record.source) <= source_limit and len(record.target) <= target_limit


def is_valid_term(term: str) -> bool:
    if len(term) < 2 or len(term) > 12:
        return False
    if SENTENCE_PUNCT_RE.search(term):
        return False
    if NON_TERM_RE.match(term):
        return False
    if not CJK_RE.search(term):
        return False
    if term.startswith(("+", "-", "/", "%")) or term.endswith(("+", "-", "/", "%")):
        return False
    return True


def category_for(term: str) -> str:
    if term in RARITY_TERMS:
        return "rarity"
    if term in RESOURCE_TERMS:
        return "resource"
    if term in STAT_TERMS or term.endswith("伤害") or term.endswith("伤害+"):
        return "stat"
    if term in ACTION_TERMS:
        return "action"
    if term in SYSTEM_TERMS:
        return "system"
    if term in OBJECT_TERMS:
        return "object"
    if term in STATUS_TERMS:
        return "status"
    if any(key in term for key in ("伤害", "攻击", "生命", "防御", "暴击")):
        return "stat"
    if any(key in term for key in ("公会", "竞技场", "战令", "签到", "商城", "商店", "基地", "防御塔", "活动")):
        return "system"
    if any(key in term for key in ("英雄", "技能", "装备", "建筑", "武器", "坐骑")):
        return "object"
    return "other"


def join_counter(counter: Counter[str], limit: int = 5) -> str:
    if not counter:
        return ""
    return " | ".join(f"{text} ({count})" for text, count in counter.most_common(limit))


def risk_for(term: str, variants: int, hits: int, suggested_en: str) -> str:
    if variants > 1 or term in HIGH_CONFUSION_TERMS or not suggested_en:
        return "high"
    if hits >= 30:
        return "medium"
    return "low"


def priority_for(risk: str, hits: int) -> str:
    if risk == "high" or hits >= 80:
        return "P1"
    if hits >= 30:
        return "P2"
    return "P3"


def note_for(
    term: str,
    variants: int,
    exact_hits: int,
    hits: int,
    suggested_en: str,
    has_actual_diff: bool,
) -> str:
    notes: list[str] = []
    if variants > 1:
        notes.append("multiple English variants detected")
    if term in ACTION_TERMS:
        notes.append("action term needs consistency review")
    if term in RARITY_TERMS:
        notes.append("rarity ladder should stay globally aligned")
    if exact_hits == 1 and hits >= 20:
        notes.append("mostly embedded usage, review with context")
    if not suggested_en:
        notes.append("no stable English match found")
    if has_actual_diff:
        notes.append("actual short usages contain manual adaptation")
    return "; ".join(notes)


def counter_to_dict(counter: Counter[str]) -> dict[str, int]:
    return {key: int(value) for key, value in sorted(counter.items()) if key}


def dict_to_counter(value: dict[str, Any] | None) -> Counter[str]:
    counter: Counter[str] = Counter()
    if not value:
        return counter
    for key, raw in value.items():
        try:
            count = int(raw)
        except (TypeError, ValueError):
            continue
        if key and count > 0:
            counter[key] = count
    return counter


def merge_counters(*counters: Counter[str]) -> Counter[str]:
    merged: Counter[str] = Counter()
    for counter in counters:
        merged.update(counter)
    return merged


def new_curated_rules() -> dict[str, Any]:
    return {"version": CURATED_VERSION, "terms": {}}


def new_observation_store() -> dict[str, Any]:
    return {"version": OBSERVATION_VERSION, "terms": {}}


def split_legacy_term_memory(memory: dict[str, Any] | None) -> tuple[dict[str, Any], dict[str, Any]]:
    curated = new_curated_rules()
    observations = new_observation_store()
    if not isinstance(memory, dict):
        return curated, observations

    for term, raw_state in memory.get("terms", {}).items():
        if not isinstance(raw_state, dict):
            continue
        curated_state = {
            "approved_en": clean_text(raw_state.get("approved_en")),
            "approved_en2": clean_text(raw_state.get("approved_en2")),
            "block_en2": bool(raw_state.get("block_en2")),
            "ignore": bool(raw_state.get("ignore")),
            "note": clean_text(raw_state.get("note")),
            "category_override": clean_text(raw_state.get("category_override")),
        }
        observation_state = {
            "observed_exact_candidates": counter_to_dict(dict_to_counter(raw_state.get("observed_exact_candidates"))),
            "observed_example_usages": counter_to_dict(dict_to_counter(raw_state.get("observed_example_usages"))),
            "observed_manual_adaptations": counter_to_dict(dict_to_counter(raw_state.get("observed_manual_adaptations"))),
            "seen_runs": max(0, int(raw_state.get("seen_runs", 0) or 0)),
            "last_seen_at": clean_text(raw_state.get("last_seen_at")),
            "last_input_digest": clean_text(raw_state.get("last_input_digest")),
        }
        if any(
            [
                curated_state["approved_en"],
                curated_state["approved_en2"],
                curated_state["block_en2"],
                curated_state["ignore"],
                curated_state["note"],
                curated_state["category_override"],
            ]
        ):
            curated["terms"][term] = curated_state
        if any(
            [
                observation_state["observed_exact_candidates"],
                observation_state["observed_example_usages"],
                observation_state["observed_manual_adaptations"],
                observation_state["seen_runs"],
                observation_state["last_seen_at"],
                observation_state["last_input_digest"],
            ]
        ):
            observations["terms"][term] = observation_state
    return curated, observations


def load_json_object(path: Path | None) -> dict[str, Any] | None:
    if path is None or not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    return payload


def legacy_experience_candidate(path: Path | None) -> Path | None:
    if path is None:
        return None
    candidate = path.with_name("term_memory.json")
    if candidate.exists():
        return candidate
    if DEFAULT_LEGACY_EXPERIENCE_STORE.exists():
        return DEFAULT_LEGACY_EXPERIENCE_STORE
    return None


def default_curated_term_state() -> dict[str, Any]:
    return {
        "approved_en": "",
        "approved_en2": "",
        "block_en2": False,
        "ignore": False,
        "note": "",
        "category_override": "",
    }


def get_curated_term_state(curated_rules: dict[str, Any], term: str, *, create: bool = True) -> dict[str, Any]:
    terms = curated_rules.setdefault("terms", {})
    if create:
        state = terms.setdefault(term, {})
    else:
        state = terms.get(term, {})
        if not isinstance(state, dict):
            state = {}
    defaults = default_curated_term_state()
    if create:
        for key, value in defaults.items():
            state.setdefault(key, value)
        return state
    defaults.update(
        {
            "approved_en": clean_text(state.get("approved_en")),
            "approved_en2": clean_text(state.get("approved_en2")),
            "block_en2": bool(state.get("block_en2")),
            "ignore": bool(state.get("ignore")),
            "note": clean_text(state.get("note")),
            "category_override": clean_text(state.get("category_override")),
        }
    )
    state = defaults
    return state


def get_observation_term_state(observations_store: dict[str, Any], term: str) -> dict[str, Any]:
    terms = observations_store.setdefault("terms", {})
    state = terms.setdefault(term, {})
    state.setdefault("observed_exact_candidates", {})
    state.setdefault("observed_manual_adaptations", {})
    state.setdefault("observed_example_usages", {})
    state.setdefault("seen_runs", 0)
    state.setdefault("last_seen_at", "")
    state.setdefault("last_input_digest", "")
    return state


def sanitize_curated_rules(payload: dict[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return new_curated_rules()
    if "terms" not in payload or not isinstance(payload["terms"], dict):
        payload = {"version": payload.get("version", CURATED_VERSION), "terms": {}}
    curated = new_curated_rules()
    curated["version"] = int(payload.get("version", CURATED_VERSION) or CURATED_VERSION)
    for term in payload["terms"]:
        if not isinstance(term, str):
            continue
        state = get_curated_term_state(curated, term)
        raw = payload["terms"].get(term)
        if isinstance(raw, dict):
            state["approved_en"] = clean_text(raw.get("approved_en"))
            state["approved_en2"] = clean_text(raw.get("approved_en2"))
            state["block_en2"] = bool(raw.get("block_en2"))
            state["ignore"] = bool(raw.get("ignore"))
            state["note"] = clean_text(raw.get("note"))
            state["category_override"] = clean_text(raw.get("category_override"))
    return curated


def sanitize_observation_store(payload: dict[str, Any] | None) -> dict[str, Any]:
    if not payload:
        return new_observation_store()
    if "terms" not in payload or not isinstance(payload["terms"], dict):
        payload = {"version": payload.get("version", OBSERVATION_VERSION), "terms": {}}
    observations = new_observation_store()
    observations["version"] = int(payload.get("version", OBSERVATION_VERSION) or OBSERVATION_VERSION)
    for term in payload["terms"]:
        if not isinstance(term, str):
            continue
        state = get_observation_term_state(observations, term)
        raw = payload["terms"].get(term)
        if isinstance(raw, dict):
            state["observed_exact_candidates"] = counter_to_dict(dict_to_counter(raw.get("observed_exact_candidates")))
            state["observed_example_usages"] = counter_to_dict(dict_to_counter(raw.get("observed_example_usages")))
            state["observed_manual_adaptations"] = counter_to_dict(dict_to_counter(raw.get("observed_manual_adaptations")))
            state["seen_runs"] = max(0, int(raw.get("seen_runs", 0) or 0))
            state["last_seen_at"] = clean_text(raw.get("last_seen_at"))
            state["last_input_digest"] = clean_text(raw.get("last_input_digest"))
    return observations


def load_curated_rules(path: Path | None) -> dict[str, Any]:
    payload = load_json_object(path)
    if payload:
        if any(
            isinstance(state, dict) and any(key.startswith("observed_") for key in state.keys())
            for state in payload.get("terms", {}).values()
        ):
            legacy_curated, _legacy_observations = split_legacy_term_memory(payload)
            return legacy_curated
        return sanitize_curated_rules(payload)

    legacy_path = legacy_experience_candidate(path)
    if legacy_path:
        legacy_payload = load_json_object(legacy_path)
        if legacy_payload:
            legacy_curated, _legacy_observations = split_legacy_term_memory(legacy_payload)
            return legacy_curated
    return new_curated_rules()


def load_observation_store(path: Path | None) -> dict[str, Any]:
    payload = load_json_object(path)
    if payload:
        if any(
            isinstance(state, dict) and any(key.startswith("approved_") or key in {"block_en2", "ignore", "note", "category_override"} for key in state.keys())
            for state in payload.get("terms", {}).values()
        ):
            _legacy_curated, legacy_observations = split_legacy_term_memory(payload)
            return legacy_observations
        return sanitize_observation_store(payload)

    legacy_path = legacy_experience_candidate(path)
    if legacy_path:
        legacy_payload = load_json_object(legacy_path)
        if legacy_payload:
            _legacy_curated, legacy_observations = split_legacy_term_memory(legacy_payload)
            return legacy_observations
    return new_observation_store()


def save_curated_rules(path: Path | None, curated_rules: dict[str, Any]) -> None:
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(sanitize_curated_rules(curated_rules), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def save_observation_store(path: Path | None, observations_store: dict[str, Any]) -> None:
    if path is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(sanitize_observation_store(observations_store), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def apply_observation_history(
    observation_state: dict[str, Any],
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> tuple[Counter[str], Counter[str], Counter[str]]:
    historical_exact = dict_to_counter(observation_state.get("observed_exact_candidates"))
    historical_examples = dict_to_counter(observation_state.get("observed_example_usages"))
    historical_manual = dict_to_counter(observation_state.get("observed_manual_adaptations"))
    return (
        merge_counters(exact_translation_counter, historical_exact),
        merge_counters(example_usage_counter, historical_examples),
        merge_counters(manual_adaptation_counter, historical_manual),
    )


def apply_curated_preferences(
    curated_state: dict[str, Any],
    term: str,
    suggested_en: str,
    example_en: str,
    en2_value: str,
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> tuple[str, str, str, Counter[str], Counter[str], Counter[str]]:
    approved_en = clean_text(curated_state.get("approved_en"))
    approved_en2 = clean_text(curated_state.get("approved_en2"))
    block_en2 = bool(curated_state.get("block_en2"))

    if approved_en:
        suggested_en = approved_en
        example_en = approved_en
    elif not example_en and exact_translation_counter:
        example_en = exact_translation_counter.most_common(1)[0][0]
        suggested_en = example_en

    if approved_en2:
        en2_value = approved_en2
    elif block_en2:
        en2_value = ""
    elif not en2_value and manual_adaptation_counter:
        en2_value = choose_en2_value(
            example_en=example_en,
            exact_diff_counter=Counter(),
            manual_counter=manual_adaptation_counter,
        )

    if curated_state.get("ignore") and term not in HIGH_CONFUSION_TERMS:
        return "", "", "", Counter(), Counter(), Counter()
    return suggested_en, example_en, en2_value, exact_translation_counter, example_usage_counter, manual_adaptation_counter


def update_observation_store(
    observation_state: dict[str, Any],
    *,
    input_digest: str,
    exact_translation_counter: Counter[str],
    example_usage_counter: Counter[str],
    manual_adaptation_counter: Counter[str],
) -> None:
    if observation_state.get("last_input_digest") == input_digest:
        observation_state["last_seen_at"] = datetime.now(timezone.utc).isoformat()
        return

    observed_exact = dict_to_counter(observation_state.get("observed_exact_candidates"))
    observed_example = dict_to_counter(observation_state.get("observed_example_usages"))
    observed_manual = dict_to_counter(observation_state.get("observed_manual_adaptations"))
    observed_exact.update(exact_translation_counter)
    observed_example.update(example_usage_counter)
    observed_manual.update(manual_adaptation_counter)

    observation_state["observed_exact_candidates"] = counter_to_dict(observed_exact)
    observation_state["observed_example_usages"] = counter_to_dict(observed_example)
    observation_state["observed_manual_adaptations"] = counter_to_dict(observed_manual)
    observation_state["seen_runs"] = int(observation_state.get("seen_runs", 0)) + 1
    observation_state["last_seen_at"] = datetime.now(timezone.utc).isoformat()
    observation_state["last_input_digest"] = input_digest


def file_digest(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def set_widths(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 60))
        worksheet.column_dimensions[letter].width = max(10, min(max_len + 2, 42))


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    set_widths(worksheet)


def resolve_column_index(headers: list[object], expected_name: str) -> int:
    normalized = {clean_text(name).lower(): index for index, name in enumerate(headers)}
    key = clean_text(expected_name).lower()
    if key not in normalized:
        available = ", ".join(str(name) for name in headers)
        raise ValueError(f"Missing column '{expected_name}'. Available headers: {available}")
    return normalized[key]


XLSX_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XLSX_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
XLSX_NS = {"a": XLSX_MAIN_NS, "rel": PACKAGE_REL_NS}


def cell_column_index(cell_ref: str) -> int:
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0
    index = 0
    for char in match.group(1):
        index = index * 26 + ord(char) - 64
    return index - 1


def xml_text(node: ET.Element | None) -> str:
    if node is None:
        return ""
    return "".join(node.itertext())


def workbook_sheet_targets(archive: ZipFile) -> list[tuple[str, str]]:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root.findall("rel:Relationship", XLSX_NS)
    }
    sheets: list[tuple[str, str]] = []
    for sheet in workbook_root.find("a:sheets", XLSX_NS).findall("a:sheet", XLSX_NS):
        rel_id = sheet.attrib[f"{{{XLSX_REL_NS}}}id"]
        target = rel_map[rel_id]
        target_path = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
        sheets.append((sheet.attrib["name"], target_path))
    return sheets


def load_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [xml_text(item) for item in root.findall("a:si", XLSX_NS)]


def raw_cell_value(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value = cell.find("a:v", XLSX_NS)
        if value is None or value.text is None:
            return ""
        try:
            return shared_strings[int(value.text)]
        except (IndexError, ValueError):
            return ""
    if cell_type == "inlineStr":
        return xml_text(cell.find("a:is", XLSX_NS))
    value = cell.find("a:v", XLSX_NS)
    return "" if value is None or value.text is None else value.text


def iter_raw_xlsx_sheets(input_path: Path) -> list[tuple[str, list[list[str]]]]:
    sheets: list[tuple[str, list[list[str]]]] = []
    with ZipFile(input_path) as archive:
        shared_strings = load_shared_strings(archive)
        for sheet_name, target_path in workbook_sheet_targets(archive):
            root = ET.fromstring(archive.read(target_path))
            rows: list[list[str]] = []
            for row in root.findall(".//a:sheetData/a:row", XLSX_NS):
                cells: list[tuple[int, str]] = []
                max_column = -1
                for cell in row.findall("a:c", XLSX_NS):
                    column_index = cell_column_index(cell.attrib.get("r", ""))
                    max_column = max(max_column, column_index)
                    cells.append((column_index, raw_cell_value(cell, shared_strings)))
                values = [""] * (max_column + 1)
                for column_index, value in cells:
                    values[column_index] = value
                rows.append(values)
            sheets.append((sheet_name, rows))
    return sheets


def records_from_rows(
    rows: list[list[object]],
    sheet_title: str,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> list[Record]:
    if not rows:
        return []
    headers = list(rows[0])
    id_index = resolve_column_index(headers, id_column)
    source_index = resolve_column_index(headers, source_column)
    target_index = None if source_only else resolve_column_index(headers, target_column)

    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        row_id = "" if id_index >= len(row_values) or row_values[id_index] is None else str(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        if not source:
            continue
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def load_records_from_raw_xlsx(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    sheets = iter_raw_xlsx_sheets(input_path)
    if not sheets:
        return [], ""
    selected_sheet: tuple[str, list[list[str]]] | None = None
    for candidate in sheets:
        if sheet_name is None or candidate[0] == sheet_name:
            selected_sheet = candidate
            break
    if selected_sheet is None:
        available = ", ".join(name for name, _rows in sheets)
        raise ValueError(f"Missing worksheet '{sheet_name}'. Available worksheets: {available}")
    title, rows = selected_sheet
    return records_from_rows(
        rows=rows,
        sheet_title=title,
        id_column=id_column,
        source_column=source_column,
        target_column=target_column,
        source_only=source_only,
    ), title


def normalized_header_lookup(headers: list[object]) -> dict[str, int]:
    return {clean_text(name).lower(): index for index, name in enumerate(headers)}


def first_matching_header(headers: list[object], candidates: list[str]) -> int | None:
    lookup = normalized_header_lookup(headers)
    for candidate in candidates:
        key = clean_text(candidate).lower()
        if key in lookup:
            return lookup[key]
    return None


def auto_records_from_sheet_rows(sheet_title: str, rows: list[list[object]]) -> list[Record]:
    if not rows:
        return []
    headers = list(rows[0])
    source_index = first_matching_header(
        headers,
        ["简体中文", "中文", "正常对话", "cn", "source", "zh", "Chinese"],
    )
    if source_index is None:
        return []
    target_index = first_matching_header(
        headers,
        ["英文", "英语", "en", "English", "优化翻译"],
    )
    id_index = first_matching_header(
        headers,
        ["唯一标识ID", "ID", "id", "章节", "关卡序号"],
    )

    records: list[Record] = []
    for row_number, row in enumerate(rows[1:], start=2):
        row_values = list(row)
        source = "" if source_index >= len(row_values) else clean_text(row_values[source_index])
        if not source:
            continue
        target = "" if target_index is None or target_index >= len(row_values) else clean_text(row_values[target_index])
        row_id = ""
        if id_index is not None and id_index < len(row_values):
            row_id = clean_text(row_values[id_index])
        if not row_id:
            row_id = f"{sheet_title}:{row_number}"
        records.append(Record(row_id=row_id, source=source, target=target))
    return records


def load_project_records(input_path: Path) -> list[Record]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        records: list[Record] = []
        for worksheet in workbook.worksheets:
            rows = list(worksheet.iter_rows(values_only=True))
            records.extend(auto_records_from_sheet_rows(worksheet.title, rows))
        workbook.close()
        return records
    except Exception:
        records = []
        for sheet_title, rows in iter_raw_xlsx_sheets(input_path):
            records.extend(auto_records_from_sheet_rows(sheet_title, rows))
        return records


def load_records(
    input_path: Path,
    sheet_name: str | None,
    id_column: str,
    source_column: str,
    target_column: str,
    source_only: bool = False,
) -> tuple[list[Record], str]:
    try:
        workbook = load_workbook(input_path, read_only=True, data_only=True)
        worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        records = records_from_rows(
            rows=rows,
            sheet_title=worksheet.title,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )
        workbook.close()
        return records, worksheet.title
    except Exception:
        return load_records_from_raw_xlsx(
            input_path=input_path,
            sheet_name=sheet_name,
            id_column=id_column,
            source_column=source_column,
            target_column=target_column,
            source_only=source_only,
        )



def build_term_rows(
    records: list[Record],
    min_hit: int,
    glossary_hit_threshold: int,
    curated_rules: dict[str, Any] | None = None,
    observations_store: dict[str, Any] | None = None,
    input_digest: str = "",
    include_empty_final_terms: bool = False,
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    curated_rules = curated_rules if curated_rules is not None else new_curated_rules()
    observations_store = observations_store if observations_store is not None else new_observation_store()
    label_counter: Counter[str] = Counter()
    label_translations: dict[str, Counter[str]] = defaultdict(Counter)

    for record in records:
        if is_valid_term(record.source):
            label_counter[record.source] += 1
            if record.target:
                label_translations[record.source][record.target] += 1

    rows_by_term: list[dict[str, object]] = []
    for term in sorted(set(label_counter)):
        hits = 0
        example_record: Record | None = None
        near_translations: Counter[str] = Counter()
        for record in records:
            if term not in record.source:
                continue
            hits += 1
            if example_record is None or len(record.source) < len(example_record.source):
                example_record = record
            if record.target and len(record.source) <= max(18, len(term) + 6):
                near_translations[record.target] += 1

        if hits < min_hit:
            continue

        exact_translations = label_translations.get(term, Counter())
        suggested_en = exact_translations.most_common(1)[0][0] if exact_translations else (
            near_translations.most_common(1)[0][0] if near_translations else ""
        )
        example_en = example_record.target if example_record and example_record.target else suggested_en

        actual_short_counter: Counter[str] = Counter()
        diff_sample: Record | None = None
        for record in records:
            if term not in record.source:
                continue
            if not is_short_usage_candidate(record=record, term=term, example_en=example_en):
                continue
            actual_short_counter[record.target] += 1
            if record.target and not is_same_or_extended_usage(example_en=example_en, actual_en=record.target):
                if diff_sample is None or (len(record.source), len(record.target)) < (len(diff_sample.source), len(diff_sample.target)):
                    diff_sample = record

        example_usage_counter, manual_adaptation_counter = split_usage_buckets(
            example_en=example_en,
            actual_counter=actual_short_counter,
        )
        exact_diff_counter = Counter(
            {
                text: count
                for text, count in exact_translations.items()
                if not is_same_or_extended_usage(example_en=example_en, actual_en=text)
            }
        )
        en2_value = choose_en2_value(
            example_en=example_en,
            exact_diff_counter=exact_diff_counter,
            manual_counter=manual_adaptation_counter,
        )

        curated_state = get_curated_term_state(curated_rules, term, create=False)
        observation_state = get_observation_term_state(observations_store, term)
        exact_translations, example_usage_counter, manual_adaptation_counter = apply_observation_history(
            observation_state=observation_state,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )
        suggested_en, example_en, en2_value, exact_translations, example_usage_counter, manual_adaptation_counter = apply_curated_preferences(
            curated_state=curated_state,
            term=term,
            suggested_en=suggested_en,
            example_en=example_en,
            en2_value=en2_value,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )
        if not suggested_en and exact_translations:
            suggested_en = exact_translations.most_common(1)[0][0]
        if not example_en:
            example_en = suggested_en
        if not suggested_en:
            suggested_en = example_en
        if not example_en and exact_translations:
            example_en = exact_translations.most_common(1)[0][0]

        update_observation_store(
            observation_state,
            input_digest=input_digest,
            exact_translation_counter=exact_translations,
            example_usage_counter=example_usage_counter,
            manual_adaptation_counter=manual_adaptation_counter,
        )

        diff_info = collect_translation_diff(example_en=example_en, actual_counter=actual_short_counter)
        risk = risk_for(term, len(exact_translations or near_translations), hits, suggested_en)
        category = clean_text(curated_state.get("category_override")) or category_for(term)
        note = note_for(
            term=term,
            variants=len(exact_translations or near_translations),
            exact_hits=label_counter[term],
            hits=hits,
            suggested_en=suggested_en,
            has_actual_diff=diff_info["has_diff"] == "Yes",
        )
        if clean_text(curated_state.get("note")):
            note = f"{note}; {clean_text(curated_state.get('note'))}" if note else clean_text(curated_state.get("note"))

        row = {
            "ID": example_record.row_id if example_record else "",
            "CN": term,
            "EN": example_en,
            "EN2": en2_value,
            "SuggestedEN": suggested_en,
            "ExactCandidates": join_counter(exact_translations or near_translations),
            "ExampleUsages": join_counter(example_usage_counter, limit=8),
            "ManualAdaptations": join_counter(manual_adaptation_counter, limit=8),
            "ActualShortUsages": join_counter(actual_short_counter, limit=8),
            "HasActualDiff": diff_info["has_diff"],
            "DiffType": diff_info["diff_type"],
            "DiffVariants": diff_info["diff_variants"],
            "SameOrFormatOnlyCount": diff_info["same_or_format_only_count"],
            "DiffCount": diff_info["diff_count"],
            "Category": category,
            "Risk": risk,
            "Priority": priority_for(risk, hits),
            "HitRows": hits,
            "ExactRows": label_counter[term],
            "ExampleID": example_record.row_id if example_record else "",
            "ExampleSource": example_record.source if example_record else "",
            "ExampleEN": example_record.target if example_record else "",
            "DiffExampleID": diff_sample.row_id if diff_sample else "",
            "DiffExampleSource": diff_sample.source if diff_sample else "",
            "DiffExampleEN": diff_sample.target if diff_sample else "",
            "Note": note,
        }
        if not curated_state.get("ignore"):
            rows_by_term.append(row)

    rows_by_term.sort(
        key=lambda row: (
            {"P1": 0, "P2": 1, "P3": 2}[row["Priority"]],
            {"high": 0, "medium": 1, "low": 2}[row["Risk"]],
            -int(row["HitRows"]),
            row["CN"],
        )
    )

    glossary_rows = [
        row for row in rows_by_term if int(row["HitRows"]) >= glossary_hit_threshold or row["Risk"] == "high"
    ]
    high_risk_rows = [row for row in rows_by_term if row["Risk"] == "high"]
    manual_rows = [row for row in rows_by_term if row["HasActualDiff"] == "Yes"]
    final_rows = list(glossary_rows) if include_empty_final_terms else [
        row for row in glossary_rows if row["EN"] or row["EN2"]
    ]
    return rows_by_term, glossary_rows, high_risk_rows, manual_rows, final_rows


def keyword_evidence(records: list[Record], keywords: set[str]) -> tuple[int, Counter[str]]:
    row_hits = 0
    keyword_counter: Counter[str] = Counter()
    for record in records:
        matched = False
        for keyword in keywords:
            if keyword in record.source:
                keyword_counter[keyword] += 1
                matched = True
        if matched:
            row_hits += 1
    return row_hits, keyword_counter


def infer_project_signals(records: list[Record], limit: int = 5) -> list[dict[str, object]]:
    signals: list[dict[str, object]] = []
    for label, keywords in PROJECT_SIGNAL_GROUPS.items():
        row_hits, evidence_counter = keyword_evidence(records, keywords)
        if not row_hits:
            continue
        signals.append(
            {
                "label": label,
                "hit_rows": row_hits,
                "evidence": join_counter(evidence_counter, limit=6),
            }
        )
    signals.sort(key=lambda item: (-int(item["hit_rows"]), str(item["label"])))
    return signals[:limit]


def markdown_table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _header in headers) + " |",
    ]
    for row in rows:
        escaped = [str(value).replace("|", "\\|") for value in row]
        lines.append("| " + " | ".join(escaped) + " |")
    return "\n".join(lines)


def source_samples(records: list[Record], signals: list[dict[str, object]], limit: int = 5) -> list[str]:
    if not signals:
        return []
    signal_labels = {str(signal["label"]) for signal in signals}
    keywords: set[str] = set()
    for label, group_keywords in PROJECT_SIGNAL_GROUPS.items():
        if label in signal_labels:
            keywords.update(group_keywords)

    samples: list[str] = []
    seen: set[str] = set()
    for record in records:
        if len(record.source) > 48:
            continue
        if not any(keyword in record.source for keyword in keywords):
            continue
        sample = f"{record.row_id}: {record.source}" if record.row_id else record.source
        if sample in seen:
            continue
        samples.append(sample)
        seen.add(sample)
        if len(samples) >= limit:
            break
    return samples


def category_distribution(rows: list[dict[str, object]]) -> Counter[str]:
    counter: Counter[str] = Counter()
    for row in rows:
        category = str(row.get("Category") or "other")
        counter[CATEGORY_LABELS.get(category, category)] += 1
    return counter


def top_terms(rows: list[dict[str, object]], limit: int = 8) -> list[str]:
    terms: list[str] = []
    for row in sorted(rows, key=lambda item: (-int(item.get("HitRows") or 0), str(item.get("CN") or "")))[:limit]:
        cn = str(row.get("CN") or "")
        en = str(row.get("EN") or "")
        en2 = str(row.get("EN2") or "")
        hit_rows = int(row.get("HitRows") or 0)
        english = en if not en2 else f"{en} / {en2}"
        terms.append(f"{cn} -> {english} ({hit_rows})" if english else f"{cn} ({hit_rows})")
    return terms


def style_guidance(signals: list[dict[str, object]], categories: Counter[str], target_coverage: int) -> list[str]:
    labels = {str(signal["label"]) for signal in signals}
    guidance = [
        "游戏内容/UI 部分尽量精简，适配移动端按钮、弹窗、任务和道具说明。",
        "剧情对话必须自然、地道、通顺，参考美剧日常对白节奏，避免逐字直译。",
        "变量、数字、换行、富文本标签和占位符必须原样保留。",
    ]
    if "合成/经营" in labels:
        guidance.append("合成、订单、生产、仓库等玩法术语保持统一，不在不同系统中来回换词。")
    if "花店/装修" in labels or "休闲/女性向" in labels:
        guidance.append("整体语气偏温暖、轻松、生活化，避免硬核、军工或过度严肃的表达。")
    if "战斗/RPG养成" in labels or categories.get("战斗属性/数值", 0):
        guidance.append("战斗、属性、技能说明优先准确表达机制，不使用夸张营销词替代数值含义。")
    if "基地/建筑经营" in labels:
        guidance.append("建筑、基地、升级线采用稳定系统名；同一建筑不要在 HQ/Base/Headquarters 之间漂移。")
    if "活动/商业化" in labels:
        guidance.append("活动、礼包、商店文案可以有吸引力，但必须短、明确、不过度夸张。")
    if "社交/公会竞争" in labels:
        guidance.append("公会、排行、竞技相关术语保持玩家社区常用表达，如 Guild、Ranking、Arena。")
    if "剧情/叙事" in labels:
        guidance.append("角色对话要保留人物关系、情绪冲突和轻喜剧节奏，可使用自然口语与缩写。")
    if target_coverage:
        guidance.append("已有英文译文视为项目历史用法；当它和术语表冲突时，优先检查是否属于手动适配 EN2。")
    return guidance


def project_type_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels and "战斗/RPG养成" in labels:
        return "科幻战机 / 飞行射击 / RPG养成"
    if "飞行/射击题材" in labels:
        return "飞行射击"
    if "战斗/RPG养成" in labels and "社交/公会竞争" in labels and "基地/建筑经营" in labels:
        return "战斗/RPG养成 / 轻SLG"
    if {"合成/经营", "花店/装修", "剧情/叙事"} <= labels:
        return "合成经营 / 花店修复 / 轻剧情休闲"
    if {"合成/经营", "剧情/叙事"} <= labels:
        return "合成经营 / 轻剧情休闲"
    if "花店/装修" in labels:
        return "花店装修 / 休闲经营"
    if "休闲/女性向" in labels:
        return "女性向休闲"
    if "战斗/RPG养成" in labels:
        return "战斗/RPG养成"
    if "剧情/叙事" in labels:
        return "剧情向休闲"
    return "移动游戏"


def target_user_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels:
        return "偏中重度、喜欢战机养成、战斗数值、装备强化和活动推进的移动端玩家。"
    if "战斗/RPG养成" in labels:
        return "偏中度、关注战力成长、英雄/装备养成、活动奖励和竞技排名的移动端玩家。"
    if "花店/装修" in labels or "休闲/女性向" in labels:
        return "偏休闲、喜欢合成/装修/经营和轻剧情的女性向或轻度玩家。"
    if "合成/经营" in labels:
        return "喜欢轻策略、收集、订单推进和长期经营成长的休闲玩家。"
    if "剧情/叙事" in labels:
        return "重视角色关系、剧情推进和自然对白体验的玩家。"
    return "移动端游戏玩家。"


def signal_hit_map(signals: list[dict[str, object]]) -> dict[str, int]:
    return {str(signal["label"]): int(signal["hit_rows"]) for signal in signals}


def content_focus_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    hits = signal_hit_map(signals)
    focus: list[str] = []
    if "飞行/射击题材" in labels:
        focus.append("战机、导弹、射击、弹幕等战斗内容")
    if "战斗/RPG养成" in labels:
        focus.append("英雄、装备、技能、属性和战力成长")
    if "基地/建筑经营" in labels:
        focus.append("建造、升级、采集、生产等基地系统")
    if hits.get("合成/经营", 0) >= 20:
        focus.append("合成、订单、生产、仓库等玩法 UI")
    if hits.get("花店/装修", 0) >= 10:
        focus.append("花店修复、装饰和生活化物件")
    if "剧情/叙事" in labels:
        focus.append("角色剧情对话")
    if "活动/商业化" in labels:
        focus.append("活动、礼包和奖励")
    return "；".join(focus) if focus else "系统 UI、玩法说明和剧情文本"


def tone_rule_from_signals(signals: list[dict[str, object]]) -> str:
    labels = {str(signal["label"]) for signal in signals}
    if "飞行/射击题材" in labels:
        return "整体语气冷静、利落、偏科幻军事；战机、装备、导弹、技能和战斗数值要专业清晰，避免可爱化、生活化或过度口语化。"
    if "战斗/RPG养成" in labels:
        return "整体语气清晰、有力量感；战斗、英雄、装备和数值成长要准确直接，避免弱化机制或夸张营销。"
    if "花店/装修" in labels or "休闲/女性向" in labels:
        return "整体语气偏轻松、温暖、生活化；涉及花店、装修、订单、合成、经营时避免硬核或过度严肃的表达。"
    if "合成/经营" in labels:
        return "整体语气轻松、清晰、偏休闲；合成、订单、生产和仓库说明要短句化，避免复杂长句。"
    return "整体语气清晰、自然、符合移动游戏语境；不要为了润色改变玩法含义。"


def build_translation_prompt(
    project_name: str,
    signals: list[dict[str, object]],
    categories: Counter[str],
    key_terms: list[str],
    target_coverage: int,
) -> str:
    project_type = project_type_from_signals(signals)
    tone_rule = tone_rule_from_signals(signals)
    term_rule = "关键术语以随附术语表为准，EN 为标准译法，EN2 为项目中稳定出现的手动适配译法。"
    if not key_terms:
        term_rule = "如未提供术语表，需先从上下文判断固定系统名，保持同一中文术语的英文一致。"
    existing_en_rule = (
        "已有英文译文代表项目历史用法；如现有译法不自然，可以优化，但不要破坏已固定的系统术语。"
        if target_coverage
        else "当前输入可能没有英文列；先按项目类型和术语表建立统一英语风格。"
    )
    return "\n".join(
        [
            f"你是一位资深游戏本地化译者，正在翻译《{project_name}》这款{project_type}游戏。",
            "译文需符合以下要求：",
            "1. 游戏内容/UI/玩法说明尽量精简，适配移动游戏按钮、弹窗、任务、道具和奖励说明；",
            "2. 剧情对话必须自然、地道、通顺，参考美剧日常对白节奏，保留角色语气、冲突、幽默和情绪，不要逐字直译；",
            f"3. {tone_rule}",
            f"4. {term_rule}",
            f"5. {existing_en_rule}",
            "6. 保留所有游戏代码、变量、数字、换行、颜色标签、HTML/富文本标签和占位符，如 {0}、%s、<color> 等；",
            "7. 无法确认的专有名词或信息缺口用 [TBD] 标记，不要自行编造设定。",
        ]
    )


def build_project_brief(
    project_name: str,
    sheet_name: str,
    records: list[Record],
    all_rows: list[dict[str, object]],
    glossary_rows: list[dict[str, object]],
    manual_rows: list[dict[str, object]],
) -> tuple[str, str]:
    source_rows = len(records)
    target_coverage = sum(1 for record in records if record.target)
    signals = infer_project_signals(records, limit=10)
    categories = category_distribution(glossary_rows or all_rows)
    key_terms = top_terms(glossary_rows or all_rows)
    project_type = project_type_from_signals(signals)
    target_user = target_user_from_signals(signals)
    content_focus = content_focus_from_signals(signals)
    tone_rule = tone_rule_from_signals(signals)
    prompt = build_translation_prompt(
        project_name=project_name,
        signals=signals,
        categories=categories,
        key_terms=key_terms,
        target_coverage=target_coverage,
    )

    markdown = "\n".join(
        [
            f"# {project_name} 翻译提示词与项目元信息",
            "",
            "## 🤖 AI 生成的专属翻译提示词",
            "",
            "```",
            prompt,
            "```",
            "",
            "## 📌 项目元信息",
            "",
            markdown_table(
                ["项目", "信息"],
                [
                    ["游戏类型", project_type],
                    ["目标用户", target_user],
                    ["内容构成", content_focus],
                    ["翻译风格", f"UI/玩法精简适配移动端；剧情自然、地道、通顺，参考美剧日常对白；{tone_rule}"],
                    ["语言资产", f"{source_rows} 条文本，已有英文 {target_coverage} 条。"],
                    ["生成日期", datetime.now().strftime("%Y-%m-%d")],
                ],
            ),
            "",
        ]
    )
    return markdown, prompt


def write_text_output(output_path: Path, content: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(content, encoding="utf-8")


def append_rows(worksheet, headers: list[str], rows: list[dict[str, object]]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_sheet(worksheet)


def write_detail_workbook(
    output_path: Path,
    sheet_name: str,
    records: list[Record],
    all_rows: list[dict[str, object]],
    glossary_rows: list[dict[str, object]],
    high_risk_rows: list[dict[str, object]],
    manual_rows: list[dict[str, object]],
    curated_rules_path: Path | None,
    observations_store_path: Path | None,
) -> None:
    workbook = Workbook()
    headers = [
        "ID",
        "CN",
        "EN",
        "EN2",
        "SuggestedEN",
        "ExactCandidates",
        "ExampleUsages",
        "ManualAdaptations",
        "ActualShortUsages",
        "HasActualDiff",
        "DiffType",
        "DiffVariants",
        "SameOrFormatOnlyCount",
        "DiffCount",
        "Category",
        "Risk",
        "Priority",
        "HitRows",
        "ExactRows",
        "ExampleID",
        "ExampleSource",
        "ExampleEN",
        "DiffExampleID",
        "DiffExampleSource",
        "DiffExampleEN",
        "Note",
    ]

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    append_rows(glossary_sheet, headers, glossary_rows)

    high_risk_sheet = workbook.create_sheet("HighRisk")
    append_rows(high_risk_sheet, headers, high_risk_rows)

    manual_sheet = workbook.create_sheet("ManualAdaptation")
    append_rows(manual_sheet, headers, manual_rows)

    all_sheet = workbook.create_sheet("Candidates")
    append_rows(all_sheet, headers, all_rows)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("SourceRows", len(records)),
        ("Sheet", sheet_name),
        ("CandidateTerms", len(all_rows)),
        ("GlossaryRows", len(glossary_rows)),
        ("HighRiskRows", len(high_risk_rows)),
        ("ManualAdaptationRows", len(manual_rows)),
        ("CuratedRules", str(curated_rules_path) if curated_rules_path else ""),
        ("ObservationsStore", str(observations_store_path) if observations_store_path else ""),
        ("Rule", "Extract short source terms from the source column and use target column only for English alignment and drift checks."),
        ("ManualAdaptation", "A term is marked as manual adaptation when short target usages introduce a stable wording different from the example EN."),
        ("LearningModel", "Curated rules keep approved EN/EN2 decisions; observation store accumulates seen variants and usage drift."),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def write_final_workbook(output_path: Path, final_rows: list[dict[str, object]]) -> None:
    workbook = Workbook()

    glossary_sheet = workbook.active
    glossary_sheet.title = "Glossary"
    final_headers = ["ID", "CN", "EN", "EN2"]
    glossary_sheet.append(final_headers)
    for row in final_rows:
        glossary_sheet.append([row.get(header, "") for header in final_headers])
    style_sheet(glossary_sheet)

    detail_sheet = workbook.create_sheet("Buckets")
    detail_headers = ["ID", "CN", "EN", "EN2", "ExampleUsages", "ManualAdaptations", "Note"]
    detail_sheet.append(detail_headers)
    for row in final_rows:
        detail_sheet.append([row.get(header, "") for header in detail_headers])
    style_sheet(detail_sheet)

    notes_sheet = workbook.create_sheet("Notes")
    notes_sheet.append(["Item", "Value"])
    for item, value in [
        ("Columns", "ID = text id, CN = source term, EN = example English, EN2 = manual adaptation English"),
        ("Rule", "EN2 remains blank when the alternative wording is not stable enough or is explicitly blocked by curated rules."),
        ("RowCount", len(final_rows)),
    ]:
        notes_sheet.append([item, value])
    style_sheet(notes_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    workbook.close()


def default_output_paths(input_path: Path, detail_output: str | None, final_output: str | None) -> tuple[Path, Path]:
    date_suffix = datetime.now().strftime("%Y%m%d")
    detail_path = Path(detail_output) if detail_output else input_path.with_name(
        f"{input_path.stem}_glossary_details_{date_suffix}.xlsx"
    )
    final_path = Path(final_output) if final_output else input_path.with_name(
        f"{input_path.stem}_ID_CN_EN_EN2_{date_suffix}.xlsx"
    )
    return detail_path, final_path


def default_project_brief_output_path(input_path: Path, project_brief_output: str | None) -> Path:
    date_suffix = datetime.now().strftime("%Y%m%d")
    return Path(project_brief_output) if project_brief_output else input_path.with_name(
        f"{input_path.stem}_project_brief_{date_suffix}.md"
    )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Extract glossary terms from a game localization language table.")
    parser.add_argument("input_path", help="Path to the source XLSX language table.")
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the first sheet.")
    parser.add_argument("--id-column", default="ID", help="ID column header. Default: ID")
    parser.add_argument("--source-column", default="cn", help="Source text column header. Default: cn")
    parser.add_argument("--target-column", default="en", help="Target text column header. Default: en")
    parser.add_argument(
        "--source-only",
        action="store_true",
        help="Treat the input as source text only and do not require a target text column.",
    )
    parser.add_argument(
        "--include-empty-final-terms",
        action="store_true",
        help="Keep final glossary rows even when EN and EN2 are blank. Useful for source-only extraction.",
    )
    parser.add_argument("--min-hit", type=int, default=5, help="Minimum hit count to keep a candidate. Default: 5")
    parser.add_argument(
        "--glossary-hit-threshold",
        type=int,
        default=10,
        help="Minimum hit count to include a candidate in the delivery glossary unless it is high risk. Default: 10",
    )
    parser.add_argument("--output", help="Path for the detailed workbook output.")
    parser.add_argument("--final-output", help="Path for the clean delivery workbook output.")
    parser.add_argument(
        "--curated-rules",
        default=str(DEFAULT_CURATED_RULES),
        help="Path to the curated glossary rules JSON file. Default: data/experience/curated_terms.json",
    )
    parser.add_argument(
        "--observations-store",
        default=str(DEFAULT_OBSERVATIONS_STORE),
        help="Path to the observed term usage JSON file. Default: data/experience/observed_terms.json",
    )
    parser.add_argument(
        "--project-name",
        help="Project name used in the project brief. Defaults to the input file stem.",
    )
    parser.add_argument(
        "--project-brief-output",
        help="Path for the project audit Markdown output. Defaults to *_project_brief_YYYYMMDD.md.",
    )
    parser.add_argument(
        "--translation-prompt-output",
        help="Optional path for a prompt-only text output extracted from the project brief.",
    )
    parser.add_argument(
        "--no-project-brief",
        action="store_true",
        help="Disable project audit Markdown generation.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    input_path = Path(args.input_path)
    detail_output_path, final_output_path = default_output_paths(
        input_path=input_path,
        detail_output=args.output,
        final_output=args.final_output,
    )
    project_name = args.project_name or input_path.stem
    project_brief_output_path = default_project_brief_output_path(
        input_path=input_path,
        project_brief_output=args.project_brief_output,
    )
    translation_prompt_output_path = Path(args.translation_prompt_output) if args.translation_prompt_output else None
    curated_rules_path = Path(args.curated_rules) if args.curated_rules else None
    observations_store_path = Path(args.observations_store) if args.observations_store else None
    curated_rules = load_curated_rules(curated_rules_path)
    observations_store = load_observation_store(observations_store_path)
    digest = file_digest(input_path)

    records, sheet_name = load_records(
        input_path=input_path,
        sheet_name=args.sheet,
        id_column=args.id_column,
        source_column=args.source_column,
        target_column=args.target_column,
        source_only=args.source_only,
    )
    all_rows, glossary_rows, high_risk_rows, manual_rows, final_rows = build_term_rows(
        records=records,
        min_hit=args.min_hit,
        glossary_hit_threshold=args.glossary_hit_threshold,
        curated_rules=curated_rules,
        observations_store=observations_store,
        input_digest=digest,
        include_empty_final_terms=args.include_empty_final_terms,
    )

    write_detail_workbook(
        output_path=detail_output_path,
        sheet_name=sheet_name,
        records=records,
        all_rows=all_rows,
        glossary_rows=glossary_rows,
        high_risk_rows=high_risk_rows,
        manual_rows=manual_rows,
        curated_rules_path=curated_rules_path,
        observations_store_path=observations_store_path,
    )
    write_final_workbook(output_path=final_output_path, final_rows=final_rows)
    project_records = records if args.no_project_brief and translation_prompt_output_path is None else (
        load_project_records(input_path) or records
    )
    project_brief_markdown, translation_prompt = build_project_brief(
        project_name=project_name,
        sheet_name=sheet_name,
        records=project_records,
        all_rows=all_rows,
        glossary_rows=glossary_rows,
        manual_rows=manual_rows,
    )
    if not args.no_project_brief:
        write_text_output(project_brief_output_path, project_brief_markdown)
    if translation_prompt_output_path is not None:
        write_text_output(translation_prompt_output_path, translation_prompt)
    save_curated_rules(curated_rules_path, curated_rules)
    save_observation_store(observations_store_path, observations_store)

    print(f"INPUT={input_path}")
    print(f"DETAIL_OUTPUT={detail_output_path}")
    print(f"FINAL_OUTPUT={final_output_path}")
    print(f"PROJECT_BRIEF_OUTPUT={project_brief_output_path if not args.no_project_brief else 'disabled'}")
    print(f"TRANSLATION_PROMPT_OUTPUT={translation_prompt_output_path or 'disabled'}")
    print(f"CURATED_RULES={curated_rules_path or 'disabled'}")
    print(f"OBSERVATIONS_STORE={observations_store_path or 'disabled'}")
    print(f"SHEET={sheet_name}")
    print(f"RECORDS={len(records)}")
    print(f"CANDIDATES={len(all_rows)}")
    print(f"GLOSSARY_ROWS={len(glossary_rows)}")
    print(f"HIGH_RISK_ROWS={len(high_risk_rows)}")
    print(f"MANUAL_ADAPTATION_ROWS={len(manual_rows)}")
    print(f"FINAL_ROWS={len(final_rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
