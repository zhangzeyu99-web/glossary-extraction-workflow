from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
from pathlib import Path
import unittest

from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = ROOT / "scripts" / "extract_glossary.py"
SPEC = importlib.util.spec_from_file_location("extract_glossary", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


class UtilityTests(unittest.TestCase):
    def test_normalize_english_ignores_case_spacing_and_plural(self):
        self.assertEqual(
            MODULE.normalize_english_for_compare("Rewards"),
            MODULE.normalize_english_for_compare("reward"),
        )
        self.assertEqual(
            MODULE.normalize_english_for_compare("Dual Guns"),
            MODULE.normalize_english_for_compare("dual   guns"),
        )

    def test_collect_translation_diff_marks_manual_adaptation(self):
        counter = MODULE.Counter(
            {
                "Sign Up": 3,
                "Registration": 2,
                "Registration Countdown": 1,
            }
        )
        diff = MODULE.collect_translation_diff("Sign Up", counter)
        self.assertEqual(diff["has_diff"], "Yes")
        self.assertEqual(diff["same_or_format_only_count"], 3)
        self.assertEqual(diff["diff_count"], 3)
        self.assertEqual(diff["diff_variants"], "Registration (2) | Registration Countdown (1)")

    def test_collect_translation_diff_ignores_context_extension(self):
        counter = MODULE.Counter(
            {
                "Registration": 2,
                "Registration Countdown": 2,
                "Registration Requirements": 1,
            }
        )
        diff = MODULE.collect_translation_diff("Registration", counter)
        self.assertEqual(diff["has_diff"], "No")
        self.assertEqual(diff["same_or_format_only_count"], 5)
        self.assertEqual(diff["diff_count"], 0)

    def test_split_usage_buckets_separates_example_and_manual_adaptation(self):
        counter = MODULE.Counter(
            {
                "Registration": 2,
                "Registration Countdown": 2,
                "Sign Up": 1,
                "Registered": 1,
            }
        )
        example_counter, manual_counter = MODULE.split_usage_buckets("Registration", counter)
        self.assertEqual(example_counter["Registration"], 2)
        self.assertEqual(example_counter["Registration Countdown"], 2)
        self.assertEqual(manual_counter["Sign Up"], 1)
        self.assertEqual(manual_counter["Registered"], 1)

    def test_choose_en2_prefers_exact_and_can_derive_compact_variant(self):
        self.assertEqual(
            MODULE.choose_en2_value(
                example_en="Registration",
                exact_diff_counter=MODULE.Counter({"Sign Up": 1}),
                manual_counter=MODULE.Counter({"Registered": 1}),
            ),
            "Sign Up",
        )
        self.assertEqual(
            MODULE.choose_en2_value(
                example_en="Level Up",
                exact_diff_counter=MODULE.Counter(),
                manual_counter=MODULE.Counter(
                    {
                        "Upgrade Module": 1,
                        "Upgrade Gold Mine": 1,
                        "Upgrade Defense Tower": 1,
                        "Upgrade Camp": 1,
                    }
                ),
            ),
            "Upgrade",
        )

    def test_build_project_brief_infers_project_signals_and_prompt(self):
        records = [
            MODULE.Record("1", "合成花束完成订单", "Merge bouquets to complete orders"),
            MODULE.Record("2", "修复花店装饰", "Restore the flower shop decor"),
            MODULE.Record("3", "领取奖励", "Claim Rewards"),
            MODULE.Record("4", "先生，我最后再问一次……您确定要这么做吗？", "Sir, I'll ask one last time... Are you sure?"),
            MODULE.Record("5", "这里唯一危险的东西是你和你的电锯！", "The only dangerous thing here is you and your chainsaw!"),
        ]
        all_rows, glossary_rows, _high_risk_rows, manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="brief-fixture",
        )

        markdown, prompt = MODULE.build_project_brief(
            project_name="Fixture Game",
            sheet_name="Sheet0",
            records=records,
            all_rows=all_rows,
            glossary_rows=glossary_rows,
            manual_rows=manual_rows,
        )

        self.assertIn("Fixture Game", markdown)
        self.assertIn("AI 生成的专属翻译提示词", markdown)
        self.assertIn("项目元信息", markdown)
        self.assertIn("合成经营", markdown)
        self.assertNotIn("输入快照", markdown)
        self.assertIn("译文需符合以下要求", prompt)
        self.assertIn("美剧日常对白", prompt)
        self.assertIn("游戏内容/UI/玩法说明尽量精简", prompt)

    def test_load_records_falls_back_to_raw_xlsx_reader(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            input_path = Path(temp_dir) / "styled_language_table.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Main"
            worksheet.append(["ID", "简体中文", "英文"])
            worksheet.append(["UIMail101", "领取", "Claim"])
            workbook.save(input_path)
            workbook.close()

            original_loader = MODULE.load_workbook

            def failing_loader(*_args, **_kwargs):
                raise TypeError("expected <class 'openpyxl.styles.fills.Fill'>")

            try:
                MODULE.load_workbook = failing_loader
                records, sheet_name = MODULE.load_records(
                    input_path=input_path,
                    sheet_name="Main",
                    id_column="ID",
                    source_column="简体中文",
                    target_column="英文",
                )
            finally:
                MODULE.load_workbook = original_loader

            self.assertEqual(sheet_name, "Main")
            self.assertEqual(records, [MODULE.Record("UIMail101", "领取", "Claim")])

    def test_project_brief_prioritizes_aircraft_combat_over_noise(self):
        records = [
            MODULE.Record("1", "战机攻击提升", "Aircraft ATK Up"),
            MODULE.Record("2", "导弹伤害增加", "Missile DMG Up"),
            MODULE.Record("3", "弹幕射击技能", "Barrage Skill"),
            MODULE.Record("4", "英雄装备强化", "Enhance Hero Gear"),
            MODULE.Record("5", "礼包奖励", "Pack Rewards"),
            MODULE.Record("6", "修复失败", "Repair failed"),
        ]
        all_rows, glossary_rows, _high_risk_rows, manual_rows, _final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=MODULE.new_curated_rules(),
            observations_store=MODULE.new_observation_store(),
            input_digest="aircraft-brief-fixture",
        )

        markdown, prompt = MODULE.build_project_brief(
            project_name="Aircraft",
            sheet_name="Sheet0",
            records=records,
            all_rows=all_rows,
            glossary_rows=glossary_rows,
            manual_rows=manual_rows,
        )

        self.assertIn("科幻战机 / 飞行射击 / RPG养成", markdown)
        self.assertIn("偏科幻军事", prompt)
        self.assertIn("避免可爱化", prompt)
        self.assertNotIn("花店修复", markdown)

    def test_project_brief_uses_extra_materials_and_notes(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            design_doc = temp_path / "setting.md"
            design_doc.write_text("项目设定：科幻战机、导弹、弹幕射击、英雄装备养成。", encoding="utf-8")
            screenshot = temp_path / "aircraft_missile_battle_ui.png"
            screenshot.write_bytes(b"fake-image")

            material_records, material_sources = MODULE.load_project_material_records(
                material_paths=[design_doc, screenshot],
                notes=["截图显示深色科幻机库和战机强化界面。"],
            )
            records = [MODULE.Record("1", "领取", "Claim")] + material_records
            markdown, prompt = MODULE.build_project_brief(
                project_name="Fixture Game",
                sheet_name="Sheet0",
                records=records,
                all_rows=[],
                glossary_rows=[],
                manual_rows=[],
                material_sources=material_sources,
            )

            self.assertIn("科幻战机 / 飞行射击 / RPG养成", markdown)
            self.assertIn("信息来源", markdown)
            self.assertIn("setting.md", markdown)
            self.assertIn("aircraft_missile_battle_ui.png", markdown)
            self.assertIn("偏科幻军事", prompt)


class MemoryTests(unittest.TestCase):
    def test_preferences_can_block_en2_and_accumulate_observations(self):
        curated = {
            "version": 1,
            "terms": {
                "奖励": {
                    "approved_en": "Reward",
                    "approved_en2": "",
                    "block_en2": True,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        observations = MODULE.new_observation_store()
        records = [
            MODULE.Record("1", "奖励", "Reward"),
            MODULE.Record("2", "奖励补发", "Promo"),
            MODULE.Record("3", "奖励", "Rewards"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=observations,
            input_digest="fixture-1",
        )
        row = {item["CN"]: item for item in final_rows}["奖励"]
        self.assertEqual(row["EN"], "Reward")
        self.assertEqual(row["EN2"], "")
        state = observations["terms"]["奖励"]
        self.assertEqual(state["seen_runs"], 1)
        self.assertIn("Reward", state["observed_exact_candidates"])

    def test_curated_rules_can_override_en_and_en2(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "Sign Up",
                    "block_en2": False,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        records = [
            MODULE.Record("1", "报名", "Sign Up"),
            MODULE.Record("2", "报名倒计时", "Registration Countdown"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=MODULE.new_observation_store(),
            input_digest="fixture-2",
        )
        row = {item["CN"]: item for item in final_rows}["报名"]
        self.assertEqual(row["EN"], "Registration")
        self.assertEqual(row["EN2"], "Sign Up")

    def test_curated_and_observation_stores_roundtrip(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            curated = MODULE.new_curated_rules()
            curated["terms"]["传说"] = {
                "approved_en": "Legend",
                "approved_en2": "Legendary",
                "block_en2": False,
                "ignore": False,
                "note": "fixture"
            }
            observations = MODULE.new_observation_store()
            observations["terms"]["传说"] = {
                "observed_exact_candidates": {"Legend": 2},
                "observed_example_usages": {"Legendary Hero": 3},
                "observed_manual_adaptations": {"Legendary": 1},
                "seen_runs": 2,
                "last_seen_at": "2026-04-24T00:00:00+00:00",
                "last_input_digest": "abc"
            }

            MODULE.save_curated_rules(curated_path, curated)
            MODULE.save_observation_store(observations_path, observations)

            loaded_curated = MODULE.load_curated_rules(curated_path)
            loaded_observations = MODULE.load_observation_store(observations_path)

            self.assertEqual(loaded_curated["terms"]["传说"]["approved_en"], "Legend")
            self.assertEqual(loaded_observations["terms"]["传说"]["seen_runs"], 2)

    def test_legacy_term_memory_can_split_into_two_layers(self):
        legacy = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "Sign Up",
                    "block_en2": False,
                    "ignore": False,
                    "note": "legacy",
                    "observed_exact_candidates": {"Registration": 2},
                    "observed_example_usages": {"Registration Countdown": 1},
                    "observed_manual_adaptations": {"Sign Up": 1},
                    "seen_runs": 2,
                    "last_seen_at": "2026-04-24T00:00:00+00:00",
                    "last_input_digest": "legacy-digest",
                }
            },
        }
        curated, observations = MODULE.split_legacy_term_memory(legacy)
        self.assertEqual(curated["terms"]["报名"]["approved_en"], "Registration")
        self.assertEqual(observations["terms"]["报名"]["seen_runs"], 2)
        self.assertIn("Sign Up", observations["terms"]["报名"]["observed_manual_adaptations"])

    def test_observations_can_backfill_en2_when_curated_en2_is_blank(self):
        curated = {
            "version": 1,
            "terms": {
                "报名": {
                    "approved_en": "Registration",
                    "approved_en2": "",
                    "block_en2": False,
                    "ignore": False,
                    "note": ""
                }
            },
        }
        observations = MODULE.new_observation_store()
        observations["terms"]["报名"] = {
            "observed_exact_candidates": {"Registration": 2},
            "observed_example_usages": {"Registration Countdown": 2},
            "observed_manual_adaptations": {"Sign Up": 4},
            "seen_runs": 2,
            "last_seen_at": "2026-04-24T00:00:00+00:00",
            "last_input_digest": "legacy-run",
        }
        records = [
            MODULE.Record("1", "报名", "Registration"),
            MODULE.Record("2", "报名条件", "Registration Requirements"),
        ]
        _all_rows, _glossary_rows, _high_risk_rows, _manual_rows, final_rows = MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=observations,
            input_digest="fixture-3",
        )
        row = {item["CN"]: item for item in final_rows}["报名"]
        self.assertEqual(row["EN"], "Registration")
        self.assertEqual(row["EN2"], "Sign Up")

    def test_build_term_rows_does_not_add_empty_curated_rules(self):
        curated = MODULE.new_curated_rules()
        records = [
            MODULE.Record("1", "奖励", "Reward"),
            MODULE.Record("2", "升级", "Level Up"),
        ]
        MODULE.build_term_rows(
            records=records,
            min_hit=1,
            glossary_hit_threshold=1,
            curated_rules=curated,
            observations_store=MODULE.new_observation_store(),
            input_digest="fixture-no-curated-pollution",
        )
        self.assertEqual(curated["terms"], {})


class CliIntegrationTests(unittest.TestCase):
    def test_cli_can_generate_source_only_final_terms(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "source_only_language_table.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "cn"])
            worksheet.append(["1", "奖励"])
            worksheet.append(["2", "奖励补发"])
            worksheet.append(["3", "升级"])
            workbook.save(input_path)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--source-only",
                    "--include-empty-final-terms",
                    "--min-hit",
                    "1",
                    "--glossary-hit-threshold",
                    "1",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            final_workbook = load_workbook(final_path, read_only=True, data_only=True)
            rows = list(final_workbook["Glossary"].iter_rows(values_only=True))
            lookup = {row[1]: row for row in rows[1:]}
            self.assertIn("奖励", lookup)
            self.assertEqual(lookup["奖励"][2], None)
            self.assertEqual(lookup["奖励"][3], None)
            final_workbook.close()

    def test_cli_generates_detail_final_and_store_outputs(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            input_path = temp_path / "sample_language_table.xlsx"
            detail_path = temp_path / "detail.xlsx"
            final_path = temp_path / "final.xlsx"
            curated_path = temp_path / "curated.json"
            observations_path = temp_path / "observations.json"
            project_brief_path = temp_path / "project_brief.md"
            prompt_path = temp_path / "translation_prompt.txt"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Sheet0"
            worksheet.append(["ID", "cn", "en"])
            worksheet.append(["1", "报名", "Registration"])
            worksheet.append(["2", "报名条件", "Registration Requirements"])
            worksheet.append(["3", "报名", "Sign Up"])
            worksheet.append(["4", "升级", "Level Up"])
            worksheet.append(["5", "升级模块", "Upgrade Module"])
            workbook.save(input_path)
            workbook.close()

            curated_path.write_text(
                json.dumps(
                    {
                        "version": 1,
                        "terms": {
                            "报名": {
                                "approved_en": "Registration",
                                "approved_en2": "Sign Up",
                                "block_en2": False,
                                "ignore": False,
                                "note": ""
                            },
                            "升级": {
                                "approved_en": "Level Up",
                                "approved_en2": "Upgrade",
                                "block_en2": False,
                                "ignore": False,
                                "note": ""
                            }
                        },
                    },
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )

            result = subprocess.run(
                [
                    sys.executable,
                    str(SCRIPT_PATH),
                    str(input_path),
                    "--output",
                    str(detail_path),
                    "--final-output",
                    str(final_path),
                    "--curated-rules",
                    str(curated_path),
                    "--observations-store",
                    str(observations_path),
                    "--project-name",
                    "Fixture Game",
                    "--project-brief-output",
                    str(project_brief_path),
                    "--translation-prompt-output",
                    str(prompt_path),
                    "--min-hit",
                    "1",
                    "--glossary-hit-threshold",
                    "1",
                ],
                cwd=ROOT,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(result.returncode, 0, msg=result.stderr or result.stdout)
            self.assertTrue(detail_path.exists())
            self.assertTrue(final_path.exists())
            self.assertTrue(curated_path.exists())
            self.assertTrue(observations_path.exists())
            self.assertTrue(project_brief_path.exists())
            self.assertTrue(prompt_path.exists())
            project_brief = project_brief_path.read_text(encoding="utf-8")
            prompt = prompt_path.read_text(encoding="utf-8")
            self.assertIn("Fixture Game", project_brief)
            self.assertIn("AI 生成的专属翻译提示词", project_brief)
            self.assertIn("项目元信息", project_brief)
            self.assertIn("译文需符合以下要求", prompt)
            self.assertIn("PROJECT_BRIEF_OUTPUT=", result.stdout)
            self.assertIn("TRANSLATION_PROMPT_OUTPUT=", result.stdout)
            self.assertIn("PROJECT_MATERIALS=0", result.stdout)

            final_workbook = load_workbook(final_path, read_only=True, data_only=True)
            glossary_sheet = final_workbook["Glossary"]
            rows = list(glossary_sheet.iter_rows(values_only=True))
            lookup = {row[1]: row for row in rows[1:]}
            self.assertEqual(lookup["报名"][2], "Registration")
            self.assertEqual(lookup["报名"][3], "Sign Up")
            self.assertEqual(lookup["升级"][2], "Level Up")
            self.assertEqual(lookup["升级"][3], "Upgrade")
            final_workbook.close()


if __name__ == "__main__":
    unittest.main()
